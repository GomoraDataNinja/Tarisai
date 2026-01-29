import streamlit as st
import pandas as pd
import numpy as np
import re
import tempfile
from datetime import datetime, date
from dateutil.parser import parse as dt_parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy as pycopy

# =========================
# Page config
# =========================
st.set_page_config(page_title="Tarisai", layout="wide")

# =========================
# Theme (Wells Fargo style)
# =========================
WF_RED = "#D71E28"
WF_BG = "#F3F3F3"
WF_CARD = "#FFFFFF"
WF_TEXT = "#111111"
WF_MUTED = "#666666"
WF_BORDER = "#E6E6E6"

# =========================
# Streamlit compatibility helpers
# =========================
def safe_primary_button(label, use_container_width=False, key=None):
    try:
        return st.button(label, type="primary", use_container_width=use_container_width, key=key)
    except TypeError:
        return st.button(label, use_container_width=use_container_width, key=key)

def safe_download_button(label, data, file_name, mime, use_container_width=False, key=None):
    try:
        return st.download_button(
            label,
            data=data,
            file_name=file_name,
            mime=mime,
            use_container_width=use_container_width,
            key=key,
        )
    except TypeError:
        return st.download_button(
            label,
            data=data,
            file_name=file_name,
            mime=mime,
            key=key,
        )

def safe_tabs(labels):
    try:
        return st.tabs(labels)
    except Exception:
        # Older Streamlit fallback
        return [st.container() for _ in labels]

def status_box(label):
    try:
        return st.status(label, expanded=True)
    except Exception:
        st.info(label)
        return None

# =========================
# CSS (locks UI clarity on deploy)
# =========================
st.markdown(
    f"""
<style>
html, body, [class*="css"] {{
    background-color: {WF_BG};
    color: {WF_TEXT};
}}
.block-container {
    padding-top: 4.2rem;
    padding-bottom: 2rem;
    max-width: 1250px;
}

.tarisai-topbar {{
    background: {WF_RED};
    color: white;
    padding: 16px 18px;
    border-radius: 14px;
    margin-bottom: 16px;
    position: relative;
}}
.tarisai-brand-center {{
    text-align: center;
    font-size: 30px;
    font-weight: 900;
    letter-spacing: 0.3px;
    line-height: 1.1;
}}
.tarisai-sub-center {{
    text-align: center;
    font-size: 12px;
    opacity: 0.98;
    margin-top: 2px;
}}
.tarisai-topbar-right {{
    position: absolute;
    right: 18px;
    top: 18px;
    font-size: 12px;
    opacity: 0.98;
}}

.tarisai-card {{
    background: {WF_CARD};
    border: 1px solid {WF_BORDER};
    border-radius: 16px;
    padding: 16px 16px;
}}
.tarisai-hero {{
    background: {WF_CARD};
    border: 1px solid {WF_BORDER};
    border-radius: 16px;
    padding: 18px 18px;
    margin-bottom: 12px;
    text-align: center;
}}
.tarisai-hero h2 {{
    margin: 0;
    padding: 0;
    font-size: 28px;
    font-weight: 900;
    color: {WF_TEXT};
}}
.tarisai-hero p {{
    margin: 8px 0 0 0;
    color: {WF_MUTED};
    font-size: 13px;
}}
.tarisai-hero .welcome {{
    margin-top: 10px;
    font-size: 14px;
    color: {WF_TEXT};
    font-weight: 700;
}}

.small-muted {{
    color: {WF_MUTED};
    font-size: 12px;
}}

hr {{
    border: none;
    border-top: 1px solid {WF_BORDER};
    margin: 10px 0;
}}

div.stButton > button[kind="primary"] {{
    background: {WF_RED} !important;
    color: white !important;
    border: 1px solid {WF_RED} !important;
    border-radius: 12px !important;
    padding: 0.6rem 0.9rem !important;
    font-weight: 800 !important;
}}
div.stButton > button {{
    border-radius: 12px !important;
}}

[data-testid="stSidebar"] {{
    background: {WF_CARD} !important;
    border-right: 1px solid {WF_BORDER};
}}

[data-testid="stFileUploader"] {{
    background: {WF_CARD};
    border: 1px solid {WF_BORDER};
    border-radius: 12px;
    padding: 10px 12px;
}}
[data-testid="stFileUploader"] section {{
    border: 2px dashed {WF_BORDER} !important;
    border-radius: 12px !important;
    transition: border-color 0.2s ease !important;
}}
[data-testid="stFileUploader"] section:hover {{
    border-color: {WF_RED} !important;
}}

/* Force uploader heading visibility on deploy */
div[data-testid="stFileUploader"] label,
div[data-testid="stFileUploader"] label[data-testid="stWidgetLabel"] {{
    color: {WF_TEXT} !important;
    font-weight: 900 !important;
    font-size: 16px !important;
    opacity: 1 !important;
}}
div[data-testid="stFileUploader"] small {{
    color: {WF_MUTED} !important;
    opacity: 1 !important;
}}

/* Progress bar */
.stProgress > div > div > div > div {{
    background-color: {WF_RED} !important;
}}

/* Hover polish */
.stButton>button {{
    transition: all 0.2s ease-in-out !important;
}}
.stButton>button:hover {{
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(215, 30, 40, 0.15) !important;
}}
</style>
""",
    unsafe_allow_html=True,
)

# =========================
# Header
# =========================
st.markdown(
    f"""
<div class="tarisai-topbar">
  <div class="tarisai-brand-center">Tarisai</div>
  <div class="tarisai-sub-center">Upload. Reconcile. Download.</div>
  <div class="tarisai-topbar-right">Recon bot</div>
</div>
""",
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="tarisai-hero">
  <h2>Reconciliation, done like an analyst</h2>
  <p>Drop your Supplier, Ledger, and Template files. I will find the tables, map columns, reconcile, and build your output file.</p>
  <div class="welcome">Hello. Upload your files and press Run. I will handle the rest.</div>
</div>
""",
    unsafe_allow_html=True,
)

# =========================
# Presets
# =========================
PRESETS = {
    "Generic (Statement)": {
        "mode": "statement",
        "statement_layout": "debit_credit",
        "stmt_keywords": {
            "ref": ["reference", "ref", "transaction ref", "document", "doc", "inv", "invoice"],
            "date": ["date", "posting", "document date", "txn date"],
            "debit": ["debit", "dr"],
            "credit": ["credit", "cr"],
            "amount": ["amount", "value", "total"],
            "desc": ["description", "details", "narration", "particulars"],
            "balance": ["balance", "running balance"],
        },
        "ledger_keywords": {
            "external": ["external", "external doc", "document no", "doc no", "reference"],
            "date": ["posting", "date", "document date"],
            "amount": ["amount", "amt", "value", "lcy", "net"],
            "desc": ["description", "details", "narration"],
            "doc_type": ["type", "document type"],
        },
        "settings": {
            "flip_ledger_sign": True,
            "amount_tolerance": 0.05,
            "date_window_days": 14,
            "use_recon_format_layout": True,
            "template_has_action": True,
            "left_start_cell": "B16",
            "right_start_cell": "H16",
            "supplier_name": "SUPPLIER",
            "min_auto_confidence": 0.70,
        },
    },
    "Generic (Invoice List)": {
        "mode": "invoice",
        "supplier_keywords": {
            "invoice": ["invoice", "vendor invoice", "inv", "bill", "reference"],
            "date": ["date", "posting", "document date"],
            "amount": ["amount", "amt", "value", "total", "net"],
            "desc": ["description", "details", "narration", "particulars"],
        },
        "ledger_keywords": {
            "external": ["external", "external doc", "document no", "doc no", "reference"],
            "date": ["posting", "date", "document date"],
            "amount": ["amount", "amt", "value", "lcy", "net"],
            "desc": ["description", "details", "narration"],
            "doc_type": ["type", "document type"],
        },
        "settings": {
            "flip_ledger_sign": True,
            "amount_tolerance": 0.05,
            "date_window_days": 14,
            "use_recon_format_layout": False,
            "template_has_action": True,
            "left_start_cell": "B16",
            "right_start_cell": "H16",
            "supplier_name": "SUPPLIER",
            "min_auto_confidence": 0.70,
        },
    },
}

# =========================
# Helpers
# =========================
def reset_file(f):
    try:
        f.seek(0)
    except Exception:
        pass

def to_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def to_num(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = to_str(x).replace(",", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan

def to_date(x, dayfirst=False):
    if isinstance(x, (pd.Timestamp, np.datetime64)):
        return pd.to_datetime(x, errors="coerce")
    s = to_str(x)
    if not s:
        return pd.NaT
    try:
        return pd.to_datetime(dt_parse(s, fuzzy=True, dayfirst=dayfirst))
    except Exception:
        return pd.NaT

def round2(x):
    if pd.isna(x):
        return np.nan
    return float(round(float(x), 2))

def round0(x):
    if pd.isna(x):
        return np.nan
    return float(round(float(x), 0))

def date_diff_days(d1, d2):
    if pd.isna(d1) or pd.isna(d2):
        return 999999
    return abs((pd.to_datetime(d1).date() - pd.to_datetime(d2).date()).days)

def normalize_invoice(s: str) -> str:
    s = (s or "").upper().strip()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def normalize_text(s: str) -> str:
    s = (s or "").upper()
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace("/", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def row_non_empty_count(row):
    return int(row.notna().sum())

def looks_like_header(cell):
    if pd.isna(cell):
        return False
    s = str(cell).strip().lower()
    if not s:
        return False
    tokens = [
        "date", "invoice", "inv", "debit", "credit", "amount",
        "ref", "reference", "details", "description",
        "doc", "posting", "external", "type", "balance"
    ]
    return any(t in s for t in tokens)

def detect_best_table_in_sheet(raw: pd.DataFrame, max_scan_rows: int = 80, sheet_name: str = ""):
    if raw is None or raw.empty:
        return None, None

    scan_rows = min(max_scan_rows, len(raw))
    best_score = -1
    best_df = None
    best_header_row = None

    for r in range(scan_rows):
        header_row = raw.iloc[r]
        non_empty = row_non_empty_count(header_row)
        if non_empty < 3:
            continue

        header_hits = sum(looks_like_header(x) for x in header_row.values)
        data_block = raw.iloc[r + 1: r + 1 + 30].copy()
        if data_block.empty:
            continue

        block_counts = [row_non_empty_count(data_block.iloc[i]) for i in range(min(len(data_block), 10))]
        consistency = float(np.mean(block_counts)) if block_counts else 0.0

        score = header_hits * 6 + non_empty + consistency

        if score > best_score:
            cols = [str(x).strip() if not pd.isna(x) else "" for x in header_row.values]
            df = raw.iloc[r + 1:].copy()
            df.columns = cols
            df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
            df = df.dropna(how="all").reset_index(drop=True)

            best_score = score
            best_df = df
            best_header_row = r

    if best_df is None:
        return None, None

    meta = {
        "sheet_name": sheet_name,
        "header_row": int(best_header_row),
        "score": float(best_score),
        "rows": int(len(best_df)),
        "cols": int(len(best_df.columns)),
    }
    return best_df, meta

def _is_date_like(x):
    try:
        dt_parse(str(x), fuzzy=True)
        return True
    except Exception:
        return False

def infer_col_by_type(df: pd.DataFrame, role: str):
    best_col = ""
    best_score = -1e9

    for c in df.columns:
        s = df[c].map(to_str)
        n = df[c].map(to_num)
        d = s.map(lambda x: True if x and _is_date_like(x) else False)

        pct_date = float(d.mean())
        pct_num = float(n.notna().mean())
        avg_len = float(s.map(len).mean())
        sparsity = float(df[c].isna().mean())

        header = str(c).lower()
        score = 0.0

        if role == "date":
            score = pct_date * 10 - sparsity * 2
            if "date" in header or "posting" in header:
                score += 3
        elif role == "amount":
            score = pct_num * 10 - sparsity * 1
            if "amount" in header or "amt" in header or "value" in header:
                score += 3
            if "debit" in header or "credit" in header:
                score -= 2
        elif role == "invoice":
            score = avg_len * 0.4 + float(s.nunique(dropna=True) / max(1, len(s))) * 2
            if "invoice" in header or "inv" in header:
                score += 5
            if "doc" in header and "invoice" not in header:
                score -= 1
        elif role == "external_doc":
            score = avg_len * 0.5 + float(s.nunique(dropna=True) / max(1, len(s))) * 1.5
            if "external" in header:
                score += 6
            if "doc" in header:
                score += 2
        elif role == "description":
            score = avg_len * 0.8
            if "desc" in header or "details" in header or "narr" in header:
                score += 3

        if score > best_score:
            best_score = score
            best_col = c

    return best_col

def first_matching_col(df: pd.DataFrame, keywords):
    if df is None or df.empty:
        return ""
    if not keywords:
        return ""
    cols = list(df.columns)
    cols_l = [str(c).lower() for c in cols]
    for kw in keywords:
        kw = str(kw).strip().lower()
        if not kw:
            continue
        for i, cl in enumerate(cols_l):
            if kw in cl:
                return cols[i]
    return ""

def choose_col(df: pd.DataFrame, chosen: str, role: str, keywords=None):
    chosen = (chosen or "").strip()
    if chosen and chosen in df.columns:
        return chosen
    by_kw = first_matching_col(df, keywords or [])
    if by_kw and by_kw in df.columns:
        return by_kw
    if role in ["date", "amount", "invoice", "external_doc", "description"]:
        return infer_col_by_type(df, role)
    return ""

def tokenize_ref(s: str):
    s = normalize_text(s)
    if not s:
        return []
    return re.findall(r"[A-Z0-9]+", s)

def token_overlap(a: str, b: str):
    ta = set(tokenize_ref(a))
    tb = set(tokenize_ref(b))
    if not ta or not tb:
        return 0.0
    return float(len(ta.intersection(tb)) / max(1, len(ta.union(tb))))

# =========================
# Key extraction rules
# =========================
AE_TOKEN = re.compile(r"\bAE[A-Z]{0,3}\d{4,}\b", re.IGNORECASE)

def extract_ae_candidates(text: str):
    t = normalize_text(text)
    found = AE_TOKEN.findall(t)
    return [normalize_invoice(x) for x in found]

def ledger_invoice_key(external_doc: str, supplier_invoice_set: set):
    candidates = extract_ae_candidates(external_doc)
    candidates = [c for c in candidates if c in supplier_invoice_set]
    if not candidates:
        return ""
    t = normalize_text(external_doc)
    last_pos = -1
    winner = candidates[0]
    for c in candidates:
        pos = t.rfind(c)
        if pos > last_pos:
            last_pos = pos
            winner = c
    return winner

DOCID_TOKEN = re.compile(r"(HREINV|HRECRN)\s*0*([0-9]+)", re.IGNORECASE)

def extract_docid(x):
    if pd.isna(x):
        return ""
    s = str(x).upper()
    m = DOCID_TOKEN.search(s)
    if not m:
        return ""
    prefix = m.group(1).upper()
    num = m.group(2)
    try:
        return prefix + str(int(num))
    except Exception:
        return prefix + num.lstrip("0")

def classify_ledger_txn(doc_type: str, external_doc: str, desc: str):
    text = f"{doc_type} {external_doc} {desc}".upper()
    if "CREDIT" in text and "MEMO" in text:
        return "credit"
    if text.startswith("CASJ") or "PAYMENT" in text or "RECEIPT" in text or "BANK" in text or "EFT" in text or "RTGS" in text:
        return "payment"
    return "invoice"

# =========================
# Normalization
# =========================
def normalize_supplier_sheet(df: pd.DataFrame, sheet_name: str, colmap: dict, keywords: dict):
    inv_col = choose_col(df, colmap.get("invoice", ""), "invoice", keywords.get("invoice"))
    date_col = choose_col(df, colmap.get("date", ""), "date", keywords.get("date"))
    amt_col = choose_col(df, colmap.get("amount", ""), "amount", keywords.get("amount"))
    desc_col = choose_col(df, colmap.get("desc", ""), "description", keywords.get("desc"))

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=False)) if date_col in df.columns else pd.NaT
    out["invoice_no_raw"] = df[inv_col].map(to_str) if inv_col in df.columns else ""
    out["invoice_key"] = out["invoice_no_raw"].map(normalize_invoice)
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""
    out["amount_signed"] = df[amt_col].map(to_num) if amt_col in df.columns else np.nan

    out["amt_r2"] = out["amount_signed"].map(round2)
    out["amt_r0"] = out["amount_signed"].map(round0)

    out["sheet_name"] = sheet_name
    out["row_id"] = [f"S_{sheet_name}_{i}" for i in range(len(out))]

    out = out[out["amount_signed"].notna()]
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)

    used = {"sheet": sheet_name, "inv_col": inv_col, "date_col": date_col, "amt_col": amt_col, "desc_col": desc_col}
    return out, used

def combine_supplier_workbook(uploaded_file, colmap: dict, keywords: dict, dedupe_on=True):
    reset_file(uploaded_file)
    xls = pd.ExcelFile(uploaded_file)
    all_norm = []
    audit = []

    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sheet)
        if df is None:
            audit.append({"sheet": sheet, "kept": False, "reason": "no table"})
            continue

        if meta["rows"] < 5 or meta["cols"] < 3:
            audit.append({"sheet": sheet, "kept": False, "reason": "too small"})
            continue

        norm, used = normalize_supplier_sheet(df, sheet, colmap, keywords)
        if norm.empty:
            audit.append({"sheet": sheet, "kept": False, "reason": "normalized empty"})
            continue

        audit.append({"sheet": sheet, "kept": True, "reason": "ok", **used, "rows": len(norm)})
        all_norm.append(norm)

    audit_df = pd.DataFrame(audit)
    if not all_norm:
        return pd.DataFrame(), audit_df

    combined = pd.concat(all_norm, ignore_index=True)

    if dedupe_on:
        combined["dedupe_key"] = (
            combined["invoice_key"].fillna("") + "|" +
            combined["doc_date"].dt.strftime("%Y-%m-%d").fillna("") + "|" +
            combined["amt_r2"].fillna(0).astype(str)
        )
        combined = combined.sort_values(["doc_date"]).drop_duplicates("dedupe_key", keep="first").reset_index(drop=True)

    return combined, audit_df

def normalize_ledger(df: pd.DataFrame, colmap: dict, keywords: dict, supplier_invoice_set: set, flip_sign: bool):
    ext_col = choose_col(df, colmap.get("external", ""), "external_doc", keywords.get("external"))
    date_col = choose_col(df, colmap.get("date", ""), "date", keywords.get("date"))
    amt_col = choose_col(df, colmap.get("amount", ""), "amount", keywords.get("amount"))
    desc_col = choose_col(df, colmap.get("desc", ""), "description", keywords.get("desc"))

    doc_type_col = (colmap.get("doc_type", "") or "").strip()
    if not doc_type_col or doc_type_col not in df.columns:
        doc_type_col = first_matching_col(df, keywords.get("doc_type", []))

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=True)) if date_col in df.columns else pd.NaT
    out["external_doc_raw"] = df[ext_col].map(to_str) if ext_col in df.columns else ""
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""
    out["doc_type"] = df[doc_type_col].map(to_str) if doc_type_col and doc_type_col in df.columns else ""

    amt = df[amt_col].map(to_num) if amt_col in df.columns else np.nan
    out["amount_signed"] = (-amt) if flip_sign else amt

    out["amt_r2"] = out["amount_signed"].map(round2)
    out["amt_r0"] = out["amount_signed"].map(round0)

    out["invoice_key_extracted"] = out["external_doc_raw"].apply(lambda x: ledger_invoice_key(x, supplier_invoice_set))
    out["txn_type"] = [
        classify_ledger_txn(dt, ext, ds)
        for dt, ext, ds in zip(out["doc_type"].tolist(), out["external_doc_raw"].tolist(), out["description"].tolist())
    ]

    out["docid"] = out["external_doc_raw"].apply(extract_docid)
    out["row_id"] = [f"L_{i}" for i in range(len(out))]

    out = out[out["amount_signed"].notna()]
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)

    used = {"ext_col": ext_col, "date_col": date_col, "amt_col": amt_col, "desc_col": desc_col, "doc_type_col": doc_type_col}
    return out, used

def normalize_statement_like_supplier(df: pd.DataFrame, sheet_name: str, layout: str, colmap: dict, keywords: dict):
    date_col = choose_col(df, colmap.get("date", ""), "date", keywords.get("date"))
    ref_col = choose_col(df, colmap.get("ref", ""), "invoice", keywords.get("ref"))
    desc_col = choose_col(df, colmap.get("desc", ""), "description", keywords.get("desc"))

    bal_col = (colmap.get("balance", "") or "").strip()
    if not bal_col or bal_col not in df.columns:
        bal_col = first_matching_col(df, keywords.get("balance", []))

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=False)) if date_col in df.columns else pd.NaT
    out["reference_raw"] = df[ref_col].map(to_str) if ref_col in df.columns else ""
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""

    if layout == "amount_only":
        amt_col = choose_col(df, colmap.get("amount", ""), "amount", keywords.get("amount"))
        amt = df[amt_col].map(to_num) if amt_col in df.columns else np.nan
        out["amount_signed"] = amt
    else:
        debit_col = (colmap.get("debit", "") or "").strip()
        if not debit_col or debit_col not in df.columns:
            debit_col = first_matching_col(df, keywords.get("debit", []))

        credit_col = (colmap.get("credit", "") or "").strip()
        if not credit_col or credit_col not in df.columns:
            credit_col = first_matching_col(df, keywords.get("credit", []))

        debit = df[debit_col].map(to_num) if debit_col and debit_col in df.columns else np.nan
        credit = df[credit_col].map(to_num) if credit_col and credit_col in df.columns else np.nan

        debit = debit.fillna(0) if isinstance(debit, pd.Series) else pd.Series([0] * len(df))
        credit = credit.fillna(0) if isinstance(credit, pd.Series) else pd.Series([0] * len(df))

        out["amount_signed"] = debit - credit

    out["abs_amount"] = out["amount_signed"].abs()
    out["docid"] = out["reference_raw"].apply(extract_docid)

    if bal_col and bal_col in df.columns:
        out["balance"] = df[bal_col].map(to_num)
    else:
        out["balance"] = np.nan

    out["sheet_name"] = sheet_name
    out["row_id"] = [f"ST_{sheet_name}_{i}" for i in range(len(out))]

    out = out.dropna(how="all")
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)

    used = {"sheet": sheet_name, "layout": layout, "ref_col": ref_col, "date_col": date_col, "desc_col": desc_col, "balance_col": bal_col}
    return out, used

def combine_statement_workbook(uploaded_file, layout: str, colmap: dict, keywords: dict):
    reset_file(uploaded_file)
    xls = pd.ExcelFile(uploaded_file)
    all_norm = []
    audit = []

    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sheet)
        if df is None:
            audit.append({"sheet": sheet, "kept": False, "reason": "no table"})
            continue

        if meta["rows"] < 5 or meta["cols"] < 3:
            audit.append({"sheet": sheet, "kept": False, "reason": "too small"})
            continue

        norm, used = normalize_statement_like_supplier(df, sheet, layout, colmap, keywords)
        if norm.empty:
            audit.append({"sheet": sheet, "kept": False, "reason": "normalized empty"})
            continue

        audit.append({"sheet": sheet, "kept": True, "reason": "ok", **used, "rows": len(norm)})
        all_norm.append(norm)

    audit_df = pd.DataFrame(audit)
    if not all_norm:
        return pd.DataFrame(), audit_df

    combined = pd.concat(all_norm, ignore_index=True)
    combined = combined.dropna(subset=["doc_date"]).reset_index(drop=True)
    return combined, audit_df

# =========================
# Auto doc type detection
# =========================
def detect_supplier_doc_type(df: pd.DataFrame):
    cols = [str(c).lower() for c in df.columns]
    has_debit = any("debit" in c or c.strip() == "dr" for c in cols)
    has_credit = any("credit" in c or c.strip() == "cr" for c in cols)
    has_balance = any("balance" in c for c in cols)
    has_invoice = any("invoice" in c or "vendor invoice" in c or c.strip() == "inv" for c in cols)

    score_statement = 0.0
    score_invoice = 0.0

    if has_debit and has_credit:
        score_statement += 0.55
    if has_balance:
        score_statement += 0.25
    if any("reference" in c or c.strip() == "ref" for c in cols):
        score_statement += 0.15

    if has_invoice:
        score_invoice += 0.60
    if any("bill" in c for c in cols):
        score_invoice += 0.10
    if any("amount" in c or "total" in c for c in cols):
        score_invoice += 0.10

    if score_statement >= score_invoice:
        mode = "statement"
        layout = "debit_credit" if (has_debit and has_credit) else "amount_only"
        confidence = min(0.99, 0.50 + score_statement)
    else:
        mode = "invoice"
        layout = None
        confidence = min(0.99, 0.50 + score_invoice)

    return mode, layout, float(confidence)

# =========================
# Matching score helpers
# =========================
def score_docid_match(amount_diff_abs: float, tol: float):
    if tol <= 0:
        tol = 0.01
    if amount_diff_abs <= tol:
        return 0.98, "DocID match and totals within tolerance"
    if amount_diff_abs <= tol * 5:
        return 0.72, "DocID match but totals differ"
    return 0.60, "DocID match but totals differ a lot"

def score_invoice_match(amount_diff_abs: float, tol: float, d_diff: int, date_window: int, overlap: float):
    score = 0.94
    reason = "Invoice reference match and totals compared"

    if amount_diff_abs > tol:
        score -= min(0.25, amount_diff_abs / max(0.01, tol) * 0.05)
        reason = "Invoice reference match but totals differ"

    if d_diff > date_window:
        score -= 0.10

    if overlap < 0.10:
        score -= 0.05

    return max(0.40, float(score)), reason

def score_payment_match(d_diff: int, date_window: int, overlap: float):
    score = 0.72
    reason = "Matched by amount and date window"

    if d_diff > date_window:
        score -= 0.18
        reason = "Matched by amount but date is far"

    if overlap >= 0.20:
        score += 0.08
        reason = "Matched by amount, date, and reference tokens"

    return max(0.35, float(score)), reason

# =========================
# Reconciliation
# =========================
def reconcile_docid(statement_df: pd.DataFrame, ledger_df: pd.DataFrame, amount_tol: float):
    stx = statement_df.copy()
    ltx = ledger_df.copy()

    stx = stx[stx["docid"].fillna("").astype(str) != ""].copy()
    ltx = ltx[ltx["docid"].fillna("").astype(str) != ""].copy()

    s_group = (
        stx.groupby("docid", as_index=False)
        .agg(doc_date=("doc_date", "min"), ref=("reference_raw", "first"), details=("description", "first"), amount=("abs_amount", "sum"))
    )

    l_group = (
        ltx.groupby("docid", as_index=False)
        .agg(doc_date=("doc_date", "min"), ref=("external_doc_raw", "first"), details=("description", "first"),
             amount=("amount_signed", lambda s: float(pd.Series(s).abs().sum())))
    )
    l_group["amount"] = l_group["amount"].abs()

    merged = pd.merge(
        s_group.add_prefix("Stmt_"),
        l_group.add_prefix("Ledg_"),
        left_on="Stmt_docid",
        right_on="Ledg_docid",
        how="outer",
        indicator=True,
    )
    merged["key"] = merged["Stmt_docid"].combine_first(merged["Ledg_docid"])
    merged["variance"] = merged["Stmt_amount"] - merged["Ledg_amount"]

    def _status(r):
        if r["_merge"] == "left_only":
            return "Post in ledger"
        if r["_merge"] == "right_only":
            return "Include on statement"
        return "Matched" if abs(float(r["variance"])) <= amount_tol else "Amount mismatch"

    merged["status"] = merged.apply(_status, axis=1)

    scores = []
    reasons = []
    for _, r in merged.iterrows():
        if r["_merge"] != "both":
            scores.append(0.55)
            reasons.append("Only one side has this DocID")
            continue
        sc, rs = score_docid_match(abs(float(r["variance"])), amount_tol)
        scores.append(sc)
        reasons.append(rs)

    merged["match_score"] = scores
    merged["match_reason"] = reasons

    stmt_only = merged[merged["_merge"] == "left_only"].copy()
    ledg_only = merged[merged["_merge"] == "right_only"].copy()

    left_table = pd.DataFrame({
        "Date": ledg_only["Ledg_doc_date"],
        "Ref": ledg_only["Ledg_docid"],
        "Details": ledg_only["Ledg_details"],
        "Amount": ledg_only["Ledg_amount"],
        "Action": "Include on statement",
    })

    right_table = pd.DataFrame({
        "Date": stmt_only["Stmt_doc_date"],
        "Ref": stmt_only["Stmt_docid"],
        "Details": stmt_only["Stmt_details"],
        "Amount": stmt_only["Stmt_amount"],
        "Action": "Post in ledger",
    })

    match_detail = pd.DataFrame({
        "match_method": "docid_total",
        "status": merged["status"],
        "match_key": merged["key"],
        "statement_date": merged["Stmt_doc_date"],
        "statement_ref": merged["Stmt_docid"],
        "statement_details": merged["Stmt_details"],
        "statement_amount": merged["Stmt_amount"],
        "ledger_date": merged["Ledg_doc_date"],
        "ledger_ref": merged["Ledg_docid"],
        "ledger_details": merged["Ledg_details"],
        "ledger_amount": merged["Ledg_amount"],
        "difference_statement_minus_ledger": merged["variance"],
        "match_score": merged["match_score"],
        "match_reason": merged["match_reason"],
    })

    mismatches = match_detail[match_detail["status"] == "Amount mismatch"].copy()

    return {
        "match_detail": match_detail,
        "left_table": left_table.sort_values(["Date", "Ref"], na_position="last"),
        "right_table": right_table.sort_values(["Date", "Ref"], na_position="last"),
        "mismatches": mismatches,
        "stmt_only": stmt_only,
        "ledg_only": ledg_only,
    }

def mk_detail(srow, lrow, status, method, match_key, score, reason):
    return {
        "match_method": method,
        "status": status,
        "match_key": match_key,
        "supplier_date": None if srow is None else srow.get("doc_date"),
        "supplier_ref": "" if srow is None else srow.get("ref"),
        "supplier_details": "" if srow is None else srow.get("details"),
        "supplier_amount": np.nan if srow is None else srow.get("amount"),
        "ledger_date": None if lrow is None else lrow.get("doc_date"),
        "ledger_ref": "" if lrow is None else lrow.get("ref"),
        "ledger_details": "" if lrow is None else lrow.get("details"),
        "ledger_amount": np.nan if lrow is None else lrow.get("amount"),
        "difference_supplier_minus_ledger": np.nan if (srow is None or lrow is None) else float(srow.get("amount") - lrow.get("amount")),
        "match_score": float(score),
        "match_reason": reason,
    }

def reconcile_invoice_style(supplier: pd.DataFrame, ledger: pd.DataFrame, amount_tol: float, date_window_days: int, min_auto_conf: float):
    details = []

    s_inv = supplier[supplier["invoice_key"].fillna("") != ""].copy()
    l_inv = ledger[ledger["invoice_key_extracted"].fillna("") != ""].copy()

    s_agg = (
        s_inv.groupby("invoice_key", as_index=False)
        .agg(doc_date=("doc_date", "min"), ref=("invoice_no_raw", "first"), details=("description", "first"), amount=("amount_signed", "sum"))
    )

    l_agg = (
        l_inv.groupby("invoice_key_extracted", as_index=False)
        .agg(doc_date=("doc_date", "min"), ref=("external_doc_raw", "first"), details=("description", "first"), amount=("amount_signed", "sum"))
        .rename(columns={"invoice_key_extracted": "invoice_key"})
    )

    s_keys = set(s_agg["invoice_key"].unique().tolist())
    l_keys = set(l_agg["invoice_key"].unique().tolist())
    all_keys = sorted(list(s_keys.union(l_keys)))

    l_map = {r["invoice_key"]: r for _, r in l_agg.iterrows()}
    s_map = {r["invoice_key"]: r for _, r in s_agg.iterrows()}

    matched_supplier_keys = set()
    matched_ledger_keys = set()

    for k in all_keys:
        srow = s_map.get(k)
        lrow = l_map.get(k)

        if srow is None and lrow is not None:
            details.append(mk_detail(None, lrow, "Missing on Supplier", "invoice_key_total", k, 0.55, "Invoice exists in ledger only"))
            continue
        if lrow is None and srow is not None:
            details.append(mk_detail(srow, None, "Missing in Ledger", "invoice_key_total", k, 0.55, "Invoice exists in supplier only"))
            continue

        diff = float(srow["amount"] - lrow["amount"])
        d_diff = date_diff_days(srow["doc_date"], lrow["doc_date"])
        overlap = token_overlap(str(srow["details"]), str(lrow["details"]))

        sc, rs = score_invoice_match(abs(diff), amount_tol, d_diff, int(date_window_days), overlap)
        status = "Matched" if (abs(diff) <= amount_tol and sc >= min_auto_conf) else ("Amount mismatch" if abs(diff) > amount_tol else "Needs review")

        details.append(
            mk_detail(
                {"doc_date": srow["doc_date"], "ref": srow["ref"], "details": srow["details"], "amount": srow["amount"]},
                {"doc_date": lrow["doc_date"], "ref": lrow["ref"], "details": lrow["details"], "amount": lrow["amount"]},
                status,
                "invoice_key_total",
                k,
                sc,
                rs,
            )
        )
        matched_supplier_keys.add(k)
        matched_ledger_keys.add(k)

    s_left = supplier[~supplier["invoice_key"].isin(matched_supplier_keys)].copy()
    l_left = ledger[~ledger["invoice_key_extracted"].isin(matched_ledger_keys)].copy()

    s_pay = s_left[s_left["invoice_key"].fillna("") == ""].copy()
    l_pay = l_left[l_left["txn_type"] == "payment"].copy()

    used_l = set()
    l_groups = l_pay.groupby("amt_r0") if not l_pay.empty else None

    if l_groups is not None:
        for _, sr in s_pay.iterrows():
            amt_key = sr["amt_r0"]
            if pd.isna(amt_key) or amt_key not in l_groups.groups:
                continue

            candidates = l_groups.get_group(amt_key).copy()
            candidates = candidates[~candidates["row_id"].isin(used_l)]
            if candidates.empty:
                continue

            candidates["date_diff"] = candidates["doc_date"].apply(lambda d: date_diff_days(d, sr["doc_date"]))
            candidates = candidates[candidates["date_diff"] <= date_window_days]
            if candidates.empty:
                continue

            best = candidates.sort_values(["date_diff"]).iloc[0]

            srow = {"doc_date": sr["doc_date"], "ref": sr["invoice_no_raw"], "details": sr["description"], "amount": sr["amount_signed"]}
            lrow = {"doc_date": best["doc_date"], "ref": best["external_doc_raw"], "details": best["description"], "amount": best["amount_signed"]}

            diff = float(srow["amount"] - lrow["amount"])
            overlap = token_overlap(str(srow["ref"]) + " " + str(srow["details"]), str(lrow["ref"]) + " " + str(lrow["details"]))
            d_diff = int(best["date_diff"])

            sc, rs = score_payment_match(d_diff, int(date_window_days), overlap)
            status = "Matched" if (abs(diff) <= amount_tol and sc >= min_auto_conf) else ("Amount mismatch" if abs(diff) > amount_tol else "Needs review")

            details.append(mk_detail(srow, lrow, status, "payment_amount_date", f"PAY|{amt_key}", sc, rs))
            used_l.add(best["row_id"])

    match_detail = pd.DataFrame(details)

    missing_in_ledger = match_detail[match_detail["status"] == "Missing in Ledger"].copy()
    missing_on_supplier = match_detail[match_detail["status"] == "Missing on Supplier"].copy()
    amount_mismatch = match_detail[match_detail["status"] == "Amount mismatch"].copy()
    needs_review = match_detail[match_detail["status"] == "Needs review"].copy()

    left_table = pd.DataFrame({
        "Date": missing_on_supplier["ledger_date"],
        "Ref": missing_on_supplier["ledger_ref"],
        "Details": missing_on_supplier["ledger_details"],
        "Amount": missing_on_supplier["ledger_amount"],
        "Action": "",
    })

    right_table = pd.DataFrame({
        "Date": missing_in_ledger["supplier_date"],
        "Ref": missing_in_ledger["supplier_ref"],
        "Details": missing_in_ledger["supplier_details"],
        "Amount": missing_in_ledger["supplier_amount"],
        "Action": "",
    })

    return {
        "match_detail": match_detail,
        "left_table": left_table,
        "right_table": right_table,
        "missing_in_ledger": missing_in_ledger,
        "missing_on_supplier": missing_on_supplier,
        "amount_mismatch": amount_mismatch,
        "needs_review": needs_review,
    }

# =========================
# Export helpers
# =========================
def parse_cell(addr: str):
    addr = (addr or "").strip().upper()
    m = re.match(r"^([A-Z]+)(\d+)$", addr)
    if not m:
        raise ValueError(f"Invalid cell address: {addr}. Use format like B16.")
    col_letters, row_str = m.group(1), m.group(2)
    return int(row_str), column_index_from_string(col_letters)

def excel_safe(v):
    if v is None:
        return None
    if pd.isna(v):
        return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.datetime64):
        dtv = pd.to_datetime(v, errors="coerce")
        if pd.isna(dtv):
            return None
        return dtv.to_pydatetime()
    if isinstance(v, date) and not isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    return v

def clear_range(ws, start_row, start_col, n_rows, n_cols):
    for r in range(start_row, start_row + n_rows):
        for c in range(start_col, start_col + n_cols):
            ws.cell(r, c).value = None

def write_table(ws, start_row, start_col, df: pd.DataFrame, action_col: bool, max_rows=6000):
    if df is None or df.empty:
        return
    cols = ["Date", "Ref", "Details", "Amount"]
    if action_col:
        cols.append("Action")
    df2 = df.copy()
    if "Action" not in df2.columns:
        df2["Action"] = ""
    df2 = df2[cols]
    rows = min(len(df2), max_rows)
    for i in range(rows):
        for j, c in enumerate(cols):
            ws.cell(start_row + i, start_col + j).value = excel_safe(df2.iloc[i, j])

def write_df_full(ws, df: pd.DataFrame, max_rows=80000):
    if df is None:
        return
    ws.append(list(df.columns))
    rows = min(len(df), max_rows)
    for i in range(rows):
        ws.append([excel_safe(v) for v in list(df.iloc[i].values)])
    for c in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

def find_row_by_label(ws, label, label_col=2):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, label_col).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None

def export_pack_recon_format(template_upload, left_df, right_df, stmt_balance, ledg_balance, as_at_dt, supplier_name, start_row=9, totals_row=29):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    reset_file(template_upload)
    with open(tmp_path, "wb") as f:
        f.write(template_upload.read())

    wb = load_workbook(tmp_path)
    ws = wb[wb.sheetnames[0]]

    ws["F2"].value = supplier_name
    ws["K5"].value = as_at_dt
    ws["K5"].number_format = "dd/mm/yyyy"

    needed_rows = max(len(left_df), len(right_df), 1)
    available_rows = totals_row - start_row
    insert_n = max(0, needed_rows - available_rows)

    if insert_n > 0:
        ws.insert_rows(totals_row, amount=insert_n)
        totals_row += insert_n

    template_style_row = start_row
    for r in range(start_row + available_rows, start_row + needed_rows):
        for c in range(2, 13):
            src = ws.cell(template_style_row, c)
            dst = ws.cell(r, c)
            dst._style = pycopy(src._style)
            dst.number_format = src.number_format

    for r in range(start_row, start_row + needed_rows):
        for c in range(2, 13):
            ws.cell(r, c).value = None

    LEFT = {"Date": 2, "Ref": 3, "Details": 4, "Amount": 5, "Action": 6}
    RIGHT = {"Date": 8, "Ref": 9, "Details": 10, "Amount": 11, "Action": 12}
    date_fmt = "dd/mm/yy"

    for i in range(needed_rows):
        r = start_row + i

        if i < len(left_df):
            it = left_df.iloc[i].to_dict()
            ws.cell(r, LEFT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, LEFT["Date"]).number_format = date_fmt
            ws.cell(r, LEFT["Ref"]).value = it.get("Ref")
            ws.cell(r, LEFT["Details"]).value = it.get("Details")
            ws.cell(r, LEFT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, LEFT["Action"]).value = it.get("Action", "")

        if i < len(right_df):
            it = right_df.iloc[i].to_dict()
            ws.cell(r, RIGHT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, RIGHT["Date"]).number_format = date_fmt
            ws.cell(r, RIGHT["Ref"]).value = it.get("Ref")
            ws.cell(r, RIGHT["Details"]).value = it.get("Details")
            ws.cell(r, RIGHT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, RIGHT["Action"]).value = it.get("Action", "")

    left_sum_range = f"E{start_row}:E{start_row + needed_rows - 1}"
    right_sum_range = f"K{start_row}:K{start_row + needed_rows - 1}"
    ws[f"E{totals_row}"].value = f"=SUM({left_sum_range})"
    ws[f"K{totals_row}"].value = f"=SUM({right_sum_range})"

    r_stmt = find_row_by_label(ws, "Balance as per Supplier Statement")
    r_ledg = find_row_by_label(ws, "Balance as per Creditors Ledger")
    if r_stmt:
        ws.cell(r_stmt, 10).value = excel_safe(stmt_balance)
    if r_ledg:
        ws.cell(r_ledg, 10).value = excel_safe(ledg_balance)

    r_adj_sup = find_row_by_label(ws, "Add: Adjustments to be made by Supplier")
    r_adj_books = find_row_by_label(ws, "Add: Adjustments to be made in our Books")
    if r_adj_sup:
        ws.cell(r_adj_sup, 10).value = f"=+E{totals_row}"
    if r_adj_books:
        ws.cell(r_adj_books, 10).value = f"=+K{totals_row}"

    wb.save(tmp_path)
    return tmp_path

def export_pack_generic(template_upload, results: dict, left_start_cell: str, right_start_cell: str, action_col: bool):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    reset_file(template_upload)
    with open(tmp_path, "wb") as f:
        f.write(template_upload.read())

    wb = load_workbook(tmp_path)
    ws = wb[wb.sheetnames[0]]

    lrow, lcol = parse_cell(left_start_cell)
    rrow, rcol = parse_cell(right_start_cell)

    clear_range(ws, lrow, lcol, 7000, 5 if action_col else 4)
    clear_range(ws, rrow, rcol, 7000, 5 if action_col else 4)

    write_table(ws, lrow, lcol, results.get("left_table"), action_col=action_col)
    write_table(ws, rrow, rcol, results.get("right_table"), action_col=action_col)

    for name in ["Match_Detail", "missing_in_ledger", "missing_on_supplier", "amount_mismatch", "needs_review"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    md = wb.create_sheet("Match_Detail")
    write_df_full(md, results.get("match_detail"))

    mi = wb.create_sheet("missing_in_ledger")
    write_df_full(mi, results.get("missing_in_ledger"))

    ms = wb.create_sheet("missing_on_supplier")
    write_df_full(ms, results.get("missing_on_supplier"))

    am = wb.create_sheet("amount_mismatch")
    write_df_full(am, results.get("amount_mismatch"))

    nr = wb.create_sheet("needs_review")
    write_df_full(nr, results.get("needs_review"))

    wb.save(tmp_path)
    return tmp_path

# =========================
# UI helpers
# =========================
def mapping_select(label, options, default_val=""):
    if default_val in options:
        idx = options.index(default_val)
    else:
        idx = 0
    return st.selectbox(label, options, index=idx)

def load_best_table_from_workbook(uploaded_file):
    reset_file(uploaded_file)
    xls = pd.ExcelFile(uploaded_file)
    best_df, best_meta = None, None
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sheet)
        if df is None:
            continue
        if best_df is None or meta["score"] > best_meta["score"]:
            best_df, best_meta = df, meta
    return best_df, best_meta

def validate_files(supplier_file, ledger_file, template_file):
    errors = []

    if not supplier_file:
        errors.append("Please upload a Supplier file")
    if not ledger_file:
        errors.append("Please upload a Ledger file")
    if not template_file:
        errors.append("Please upload a Template file")

    if errors:
        return False, errors

    valid_extensions = [".xlsx", ".xls"]
    for file_obj, name in [(supplier_file, "Supplier"), (ledger_file, "Ledger"), (template_file, "Template")]:
        if file_obj and not any(file_obj.name.lower().endswith(ext) for ext in valid_extensions):
            errors.append(f"{name} file must be an Excel file (.xlsx or .xls)")

    if errors:
        return False, errors

    return True, []

def show_welcome_state():
    st.markdown(
        """
<div class="tarisai-card" style="text-align: center; padding: 40px 20px;">
    <h3 style="margin-bottom: 12px; color: #111111; font-weight: 900;">Ready to reconcile</h3>
    <p style="color: #666666; max-width: 650px; margin: 0 auto;">
        Upload your Supplier, Ledger, and Template Excel files. Tarisai will detect tables, match transactions, and generate the reconciliation output.
    </p>
</div>
""",
        unsafe_allow_html=True,
    )

# =========================
# Main screen
# =========================
left, right = st.columns([2.2, 1.0])

with left:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)

    supplier_file = st.file_uploader(
        "Supplier file",
        type=["xlsx", "xls"],
        help="Upload supplier statement or invoice list",
        label_visibility="visible",
    )

    ledger_file = st.file_uploader(
        "Ledger file",
        type=["xlsx", "xls"],
        help="Upload your accounts ledger extract",
        label_visibility="visible",
    )

    template_file = st.file_uploader(
        "Recon format",
        type=["xlsx", "xls"],
        help="Upload the reconciliation template",
        label_visibility="visible",
    )

    st.markdown("</div>", unsafe_allow_html=True)

with right:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)

    preset_options = ["Generic (Statement)", "Generic (Invoice List)"]
    preset_name = st.selectbox("Preset", preset_options, index=0)

    preset = PRESETS[preset_name]
    settings = preset.get("settings", {})

    supplier_name = st.text_input("Supplier name", value=str(settings.get("supplier_name", "SUPPLIER")))

    st.markdown("<div style='margin-top: 20px; margin-bottom: 20px;'>", unsafe_allow_html=True)
    run_btn = safe_primary_button("Run reconciliation", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)
    st.markdown('<div class="small-muted">Open advanced only if the preview looks wrong.</div>', unsafe_allow_html=True)
    adv_open = st.checkbox("Open advanced", value=False)

    st.markdown("</div>", unsafe_allow_html=True)

if not (supplier_file and ledger_file and template_file):
    show_welcome_state()
    st.stop()

# =========================
# Advanced settings
# =========================
flip_ledger_sign = bool(settings.get("flip_ledger_sign", True))
amount_tolerance = float(settings.get("amount_tolerance", 0.05))
date_window_days = int(settings.get("date_window_days", 14))
use_recon_format_layout = bool(settings.get("use_recon_format_layout", True))
template_has_action = bool(settings.get("template_has_action", True))
left_start_cell = str(settings.get("left_start_cell", "B16"))
right_start_cell = str(settings.get("right_start_cell", "H16"))
min_auto_confidence = float(settings.get("min_auto_confidence", 0.70))

stmt_layout_override = None
forced_mode = None
dedupe_on = True

if adv_open:
    with st.sidebar:
        st.subheader("Advanced")
        st.caption("Use this only if the preview looks wrong")
        forced_mode = st.radio("Force supplier type", ["Auto", "Statement", "Invoice List"], index=0)
        stmt_layout_override = st.radio("Statement layout", ["Auto", "debit_credit", "amount_only"], index=0)
        flip_ledger_sign = st.checkbox("Flip ledger sign", value=flip_ledger_sign)
        amount_tolerance = st.number_input("Amount tolerance", min_value=0.0, value=float(amount_tolerance), step=0.01)
        date_window_days = st.number_input("Date window days", min_value=0, value=int(date_window_days), step=1)
        min_auto_confidence = st.number_input("Min auto match confidence", min_value=0.0, max_value=1.0, value=float(min_auto_confidence), step=0.01)

        use_recon_format_layout = st.checkbox("Template is RECON FORMAT layout", value=use_recon_format_layout)
        template_has_action = st.checkbox("Template has Action column", value=template_has_action)

        st.caption("Used when RECON FORMAT layout is off")
        left_start_cell = st.text_input("Left table start cell", value=left_start_cell)
        right_start_cell = st.text_input("Right table start cell", value=right_start_cell)

        st.divider()
        st.subheader("Invoice list options")
        dedupe_on = st.checkbox("Remove duplicates (invoice + date + amount)", value=True)

# =========================
# Run reconciliation
# =========================
if run_btn:
    is_valid, errors = validate_files(supplier_file, ledger_file, template_file)
    if not is_valid:
        for error in errors:
            st.error(error)
        st.stop()

    sbox = status_box("Tarisai is scanning your files")

    def step(msg):
        if sbox is not None:
            sbox.write(msg)
        else:
            st.info(msg)

    step("Files validated. Scanning for the main tables.")

    with st.spinner("Detecting ledger table"):
        ledger_best_df, ledger_meta = load_best_table_from_workbook(ledger_file)

    if ledger_best_df is None:
        st.error("No usable ledger table detected. Please check your ledger file format.")
        st.stop()

    step(f"Ledger table found on sheet '{ledger_meta['sheet_name']}' (header row {ledger_meta['header_row'] + 1}).")

    with st.spinner("Detecting supplier table"):
        supplier_best_df, supplier_meta = load_best_table_from_workbook(supplier_file)

    if supplier_best_df is None:
        st.error("No usable supplier table detected. Please check your supplier file format.")
        st.stop()

    step(f"Supplier table found on sheet '{supplier_meta['sheet_name']}' (header row {supplier_meta['header_row'] + 1}).")

    auto_mode, auto_stmt_layout, doc_conf = detect_supplier_doc_type(supplier_best_df)
    chosen_mode = auto_mode
    chosen_stmt_layout = auto_stmt_layout

    if forced_mode and forced_mode != "Auto":
        chosen_mode = "statement" if forced_mode == "Statement" else "invoice"
        step(f"Supplier type set to {forced_mode}.")
    else:
        label = "Statement" if chosen_mode == "statement" else "Invoice List"
        step(f"Supplier type detected as {label} (confidence {doc_conf:.0%}).")

    if chosen_mode == "statement":
        if stmt_layout_override and stmt_layout_override != "Auto":
            chosen_stmt_layout = stmt_layout_override
            step(f"Statement layout set to {chosen_stmt_layout}.")
        else:
            step(f"Statement layout set to {chosen_stmt_layout}.")

    tabs = safe_tabs(["Results", "Previews", "Audit"])
    tab_results, tab_previews, tab_audit = tabs[0], tabs[1], tabs[2]

    with tab_previews:
        st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
        st.subheader("Detected supplier table preview")
        st.write(f"Sheet: {supplier_meta['sheet_name']} | Header row: {supplier_meta['header_row'] + 1}")
        st.dataframe(supplier_best_df.head(30), use_container_width=True)

        st.subheader("Detected ledger table preview")
        st.write(f"Sheet: {ledger_meta['sheet_name']} | Header row: {ledger_meta['header_row'] + 1}")
        st.dataframe(ledger_best_df.head(30), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    if chosen_mode == "statement":
        stmt_kw = PRESETS["Generic (Statement)"]["stmt_keywords"]
        led_kw = PRESETS["Generic (Statement)"]["ledger_keywords"]

        stmt_auto = {
            "ref": first_matching_col(supplier_best_df, stmt_kw.get("ref", [])) or infer_col_by_type(supplier_best_df, "invoice"),
            "date": first_matching_col(supplier_best_df, stmt_kw.get("date", [])) or infer_col_by_type(supplier_best_df, "date"),
            "debit": first_matching_col(supplier_best_df, stmt_kw.get("debit", [])),
            "credit": first_matching_col(supplier_best_df, stmt_kw.get("credit", [])),
            "amount": first_matching_col(supplier_best_df, stmt_kw.get("amount", [])) or infer_col_by_type(supplier_best_df, "amount"),
            "desc": first_matching_col(supplier_best_df, stmt_kw.get("desc", [])) or infer_col_by_type(supplier_best_df, "description"),
            "balance": first_matching_col(supplier_best_df, stmt_kw.get("balance", [])),
        }

        ledger_auto = {
            "external": first_matching_col(ledger_best_df, led_kw.get("external", [])) or infer_col_by_type(ledger_best_df, "external_doc"),
            "date": first_matching_col(ledger_best_df, led_kw.get("date", [])) or infer_col_by_type(ledger_best_df, "date"),
            "amount": first_matching_col(ledger_best_df, led_kw.get("amount", [])) or infer_col_by_type(ledger_best_df, "amount"),
            "desc": first_matching_col(ledger_best_df, led_kw.get("desc", [])) or infer_col_by_type(ledger_best_df, "description"),
            "doc_type": first_matching_col(ledger_best_df, led_kw.get("doc_type", [])),
        }

        stmt_colmap = stmt_auto.copy()
        ledger_colmap = ledger_auto.copy()

        if adv_open:
            with st.sidebar:
                st.divider()
                st.subheader("Column mapping")

                stmt_cols = ["(auto)"] + list(supplier_best_df.columns)
                ledger_cols = ["(auto)"] + list(ledger_best_df.columns)

                st.write("Statement columns")
                stmt_colmap["date"] = mapping_select("Statement date", stmt_cols, stmt_auto["date"])
                stmt_colmap["ref"] = mapping_select("Statement reference", stmt_cols, stmt_auto["ref"])
                stmt_colmap["desc"] = mapping_select("Statement description", stmt_cols, stmt_auto["desc"])
                stmt_colmap["balance"] = mapping_select("Statement balance (optional)", stmt_cols, stmt_auto["balance"])

                if chosen_stmt_layout == "amount_only":
                    stmt_colmap["amount"] = mapping_select("Statement amount", stmt_cols, stmt_auto["amount"])
                else:
                    stmt_colmap["debit"] = mapping_select("Statement debit", stmt_cols, stmt_auto["debit"])
                    stmt_colmap["credit"] = mapping_select("Statement credit", stmt_cols, stmt_auto["credit"])

                st.write("Ledger columns")
                ledger_colmap["date"] = mapping_select("Ledger date", ledger_cols, ledger_auto["date"])
                ledger_colmap["external"] = mapping_select("Ledger external doc", ledger_cols, ledger_auto["external"])
                ledger_colmap["amount"] = mapping_select("Ledger amount", ledger_cols, ledger_auto["amount"])
                ledger_colmap["desc"] = mapping_select("Ledger description", ledger_cols, ledger_auto["desc"])
                ledger_colmap["doc_type"] = mapping_select("Ledger doc type (optional)", ledger_cols, ledger_auto["doc_type"])

        for k in list(stmt_colmap.keys()):
            if stmt_colmap.get(k) == "(auto)":
                stmt_colmap[k] = ""
        for k in list(ledger_colmap.keys()):
            if ledger_colmap.get(k) == "(auto)":
                ledger_colmap[k] = ""

        step("Column mapping complete. Normalizing data.")

        with st.spinner("Combining statement sheets"):
            stmt_combined, stmt_audit = combine_statement_workbook(
                supplier_file,
                layout=chosen_stmt_layout,
                colmap=stmt_colmap,
                keywords=stmt_kw,
            )

        if stmt_combined.empty:
            st.error("Statement normalization returned no rows. Check column mapping in advanced settings.")
            st.stop()

        step(f"Statement rows normalized: {len(stmt_combined)}")

        supplier_invoice_set = set()

        with st.spinner("Normalizing ledger"):
            ledger_norm, ledger_used = normalize_ledger(
                ledger_best_df,
                colmap=ledger_colmap,
                keywords=led_kw,
                supplier_invoice_set=supplier_invoice_set,
                flip_sign=flip_ledger_sign,
            )

        step(f"Ledger rows normalized: {len(ledger_norm)}")

        step("Reconciling by DocID totals")

        with st.spinner("Reconciling"):
            results_docid = reconcile_docid(stmt_combined, ledger_norm, amount_tol=float(amount_tolerance))

        md = results_docid["match_detail"].copy()
        hi = int((md["match_score"] >= min_auto_confidence).sum())
        mism = int((md["status"] == "Amount mismatch").sum())
        left_ct = int(len(results_docid["left_table"]))
        right_ct = int(len(results_docid["right_table"]))

        stmt_balance = float(stmt_combined["balance"].dropna().iloc[-1]) if stmt_combined["balance"].dropna().shape[0] else np.nan
        ledg_bal_col = first_matching_col(ledger_best_df, ["balance", "balance (lcy)"])
        ledg_balance = float(ledger_best_df[ledg_bal_col].map(to_num).dropna().iloc[-1]) if ledg_bal_col and ledger_best_df[ledg_bal_col].map(to_num).dropna().shape[0] else np.nan

        stmt_max_date = pd.to_datetime(stmt_combined["doc_date"], errors="coerce").max()
        ledg_max_date = pd.to_datetime(ledger_norm["doc_date"], errors="coerce").max()
        cands = [d for d in [stmt_max_date, ledg_max_date] if pd.notna(d)]
        as_at = max(cands).to_pydatetime() if cands else datetime.now()

        step("Building output file")

        with st.spinner("Writing output workbook"):
            if use_recon_format_layout:
                out_path = export_pack_recon_format(
                    template_upload=template_file,
                    left_df=results_docid["left_table"],
                    right_df=results_docid["right_table"],
                    stmt_balance=stmt_balance,
                    ledg_balance=ledg_balance,
                    as_at_dt=as_at,
                    supplier_name=supplier_name,
                )
            else:
                generic_pack = {
                    "left_table": results_docid["left_table"],
                    "right_table": results_docid["right_table"],
                    "match_detail": results_docid["match_detail"],
                    "missing_in_ledger": results_docid["stmt_only"],
                    "missing_on_supplier": results_docid["ledg_only"],
                    "amount_mismatch": results_docid["mismatches"],
                    "needs_review": results_docid["match_detail"][results_docid["match_detail"]["match_score"] < min_auto_confidence].copy(),
                }
                out_path = export_pack_generic(
                    template_upload=template_file,
                    results=generic_pack,
                    left_start_cell=left_start_cell,
                    right_start_cell=right_start_cell,
                    action_col=template_has_action,
                )

        if sbox is not None:
            try:
                sbox.update(label="Reconciliation complete", state="complete", expanded=False)
            except Exception:
                pass

        with tab_results:
            st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
            st.subheader("Reconciliation Results")

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("High confidence matches", hi)
            with c2:
                st.metric("Amount mismatches", mism)
            with c3:
                st.metric("Missing items", left_ct + right_ct)
            with c4:
                st.metric("Total groups", len(md))

            st.markdown("<hr/>", unsafe_allow_html=True)

            st.subheader("Output tables")
            a, b = st.columns(2)
            with a:
                st.write("Ledger items missing on statement")
                st.dataframe(results_docid["left_table"].head(500), use_container_width=True)
            with b:
                st.write("Statement items missing in ledger")
                st.dataframe(results_docid["right_table"].head(500), use_container_width=True)

            with st.expander("Match detail (top 2000)", expanded=False):
                st.dataframe(md.sort_values(["match_score"], ascending=False).head(2000), use_container_width=True)

            with open(out_path, "rb") as f:
                safe_download_button(
                    "Download reconciliation output",
                    data=f,
                    file_name=f"tarisai_recon_{supplier_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            st.markdown("</div>", unsafe_allow_html=True)

        with tab_audit:
            st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
            st.subheader("Processing Audit")
            st.write("Statement sheets audit")
            st.dataframe(stmt_audit, use_container_width=True)
            st.write("Ledger columns used")
            st.json(ledger_used)
            st.markdown("</div>", unsafe_allow_html=True)

    else:
        sup_kw = PRESETS["Generic (Invoice List)"]["supplier_keywords"]
        led_kw = PRESETS["Generic (Invoice List)"]["ledger_keywords"]

        sup_auto = {
            "invoice": first_matching_col(supplier_best_df, sup_kw.get("invoice", [])) or infer_col_by_type(supplier_best_df, "invoice"),
            "date": first_matching_col(supplier_best_df, sup_kw.get("date", [])) or infer_col_by_type(supplier_best_df, "date"),
            "amount": first_matching_col(supplier_best_df, sup_kw.get("amount", [])) or infer_col_by_type(supplier_best_df, "amount"),
            "desc": first_matching_col(supplier_best_df, sup_kw.get("desc", [])) or infer_col_by_type(supplier_best_df, "description"),
        }

        ledger_auto = {
            "external": first_matching_col(ledger_best_df, led_kw.get("external", [])) or infer_col_by_type(ledger_best_df, "external_doc"),
            "date": first_matching_col(ledger_best_df, led_kw.get("date", [])) or infer_col_by_type(ledger_best_df, "date"),
            "amount": first_matching_col(ledger_best_df, led_kw.get("amount", [])) or infer_col_by_type(ledger_best_df, "amount"),
            "desc": first_matching_col(ledger_best_df, led_kw.get("desc", [])) or infer_col_by_type(ledger_best_df, "description"),
            "doc_type": first_matching_col(ledger_best_df, led_kw.get("doc_type", [])),
        }

        supplier_colmap = sup_auto.copy()
        ledger_colmap = ledger_auto.copy()

        if adv_open:
            with st.sidebar:
                st.divider()
                st.subheader("Column mapping")
                supplier_cols = ["(auto)"] + list(supplier_best_df.columns)
                ledger_cols = ["(auto)"] + list(ledger_best_df.columns)

                st.write("Supplier columns")
                supplier_colmap["invoice"] = mapping_select("Supplier invoice", supplier_cols, sup_auto["invoice"])
                supplier_colmap["date"] = mapping_select("Supplier date", supplier_cols, sup_auto["date"])
                supplier_colmap["amount"] = mapping_select("Supplier amount", supplier_cols, sup_auto["amount"])
                supplier_colmap["desc"] = mapping_select("Supplier description", supplier_cols, sup_auto["desc"])

                st.write("Ledger columns")
                ledger_colmap["date"] = mapping_select("Ledger date", ledger_cols, ledger_auto["date"])
                ledger_colmap["external"] = mapping_select("Ledger external doc", ledger_cols, ledger_auto["external"])
                ledger_colmap["amount"] = mapping_select("Ledger amount", ledger_cols, ledger_auto["amount"])
                ledger_colmap["desc"] = mapping_select("Ledger description", ledger_cols, ledger_auto["desc"])
                ledger_colmap["doc_type"] = mapping_select("Ledger doc type (optional)", ledger_cols, ledger_auto["doc_type"])

        for k in list(supplier_colmap.keys()):
            if supplier_colmap.get(k) == "(auto)":
                supplier_colmap[k] = ""
        for k in list(ledger_colmap.keys()):
            if ledger_colmap.get(k) == "(auto)":
                ledger_colmap[k] = ""

        step("Column mapping complete. Combining supplier sheets.")

        with st.spinner("Combining supplier sheets"):
            supplier_combined, supplier_audit = combine_supplier_workbook(
                supplier_file,
                colmap=supplier_colmap,
                keywords=sup_kw,
                dedupe_on=dedupe_on,
            )

        if supplier_combined.empty:
            st.error("Supplier normalization returned no rows. Check column mapping in advanced settings.")
            st.stop()

        step(f"Supplier rows normalized: {len(supplier_combined)}")

        supplier_invoice_set = set(supplier_combined["invoice_key"].dropna().astype(str).unique().tolist())
        supplier_invoice_set = {x for x in supplier_invoice_set if x}

        with st.spinner("Normalizing ledger"):
            ledger_norm, ledger_used = normalize_ledger(
                ledger_best_df,
                colmap=ledger_colmap,
                keywords=led_kw,
                supplier_invoice_set=supplier_invoice_set,
                flip_sign=flip_ledger_sign,
            )

        step(f"Ledger rows normalized: {len(ledger_norm)}")

        step("Reconciling by invoice references, then payment fallback")

        with st.spinner("Reconciling"):
            results = reconcile_invoice_style(
                supplier_combined,
                ledger_norm,
                amount_tol=float(amount_tolerance),
                date_window_days=int(date_window_days),
                min_auto_conf=float(min_auto_confidence),
            )

        md = results["match_detail"].copy()
        hi = int((md["match_score"] >= min_auto_confidence).sum())
        review = int((md["status"] == "Needs review").sum())
        mism = int((md["status"] == "Amount mismatch").sum())
        left_ct = int(len(results["left_table"]))
        right_ct = int(len(results["right_table"]))

        step("Building output file")

        with st.spinner("Writing output workbook"):
            out_path = export_pack_generic(
                template_upload=template_file,
                results=results,
                left_start_cell=left_start_cell,
                right_start_cell=right_start_cell,
                action_col=template_has_action,
            )

        if sbox is not None:
            try:
                sbox.update(label="Reconciliation complete", state="complete", expanded=False)
            except Exception:
                pass

        with tab_results:
            st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
            st.subheader("Reconciliation Results")

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("High confidence matches", hi)
            with c2:
                st.metric("Needs review", review)
            with c3:
                st.metric("Amount mismatches", mism)
            with c4:
                st.metric("Missing items", left_ct + right_ct)

            st.markdown("<hr/>", unsafe_allow_html=True)

            st.subheader("Output tables")
            a, b = st.columns(2)
            with a:
                st.write("Ledger items missing on supplier")
                st.dataframe(results["left_table"].head(500), use_container_width=True)
            with b:
                st.write("Supplier items missing in ledger")
                st.dataframe(results["right_table"].head(500), use_container_width=True)

            with st.expander("Needs review (top 2000)", expanded=False):
                st.dataframe(results["needs_review"].head(2000), use_container_width=True)

            with st.expander("Match detail (top 2500)", expanded=False):
                st.dataframe(md.sort_values(["match_score"], ascending=False).head(2500), use_container_width=True)

            with open(out_path, "rb") as f:
                safe_download_button(
                    "Download reconciliation output",
                    data=f,
                    file_name=f"tarisai_recon_{supplier_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            st.markdown("</div>", unsafe_allow_html=True)

        with tab_audit:
            st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
            st.subheader("Processing Audit")
            st.write("Supplier sheets audit")
            st.dataframe(supplier_audit, use_container_width=True)
            st.write("Ledger columns used")
            st.json(ledger_used)
            st.markdown("</div>", unsafe_allow_html=True)

# =========================
# Footer
# =========================
st.markdown(
    """
<div style="text-align: center; margin-top: 40px; padding: 20px; color: #666666; font-size: 12px;">
    <hr style="margin-bottom: 20px;">
    <div>Tarisai Reconciliation Tool</div>
    <div style="margin-top: 8px;">Upload. Reconcile. Download.</div>
</div>
""",
    unsafe_allow_html=True,
)


